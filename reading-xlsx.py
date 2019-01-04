#Reading data from .xlsx file test
from openpyxl import load_workbook

#Open .xlsx file as Workbook
wb = load_workbook('data.xlsx')
#Open sheet on wb acessing it by name index
ws = wb[u'fields-data']

def read_data():
    map = {}
    data = []

	#Get column names and map to their title
    for cell in ws[1]:
        map[cell.column] = cell.value

    for row in ws.iter_rows(min_row=2, max_col=len(map.keys()), max_row=ws.max_row):
        person_data = {}
        for cell in row:
            person_data[map[cell.column]] = cell.value
        if not filter(lambda x: x is not None, person_data.values()):
			break
        data.append(person_data)

    return data

#print read_data()